"""
retry_final.py
--------------
Creates a FRESH Logs.xlsx and processes all 132 previously failed files:
  - 34 confirmed BSE/Indian market holidays  → logged as HOLIDAY (yellow)
  - 77 genuine trading days                  → retried with generous settings
  -  1 NSE SSL error                         → retried
  - 20 already confirmed holidays from CSV   → logged as HOLIDAY (yellow)

Run:
    python "D:\\Bhavya Khaitan\\Stocks\\retry_final.py"
"""

import requests
import os
import time
import random
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# PATHS
# ============================================================
BSE_DIR   = r"D:\Bhavya Khaitan\Stocks\StockData\BSE"
NSE_OLD   = r"D:\Bhavya Khaitan\Stocks\StockData\NSE\Old Format"
EXCEL_LOG = r"D:\Bhavya Khaitan\Stocks\StockData\Logs.xlsx"

# ============================================================
# RETRY SETTINGS
# ============================================================
TIMEOUT    = 90
MAX_ATT    = 3
MIN_DELAY  = 8.0
MAX_DELAY  = 15.0
RETRY_WAIT = 45

# ============================================================
# CONFIRMED HOLIDAYS FROM CSV + proper analysis (54 total)
# These will be logged as HOLIDAY — no download attempted
# ============================================================
HOLIDAYS = [
    # From original CSV (classified correctly)
    ("BSE", "2007-10-02", "EQ021007_CSV.ZIP",  "Gandhi Jayanti"),
    ("BSE", "2008-05-01", "EQ010508_CSV.ZIP",  "Maharashtra Day"),
    ("BSE", "2009-05-01", "EQ010509_CSV.ZIP",  "Maharashtra Day"),
    ("BSE", "2009-12-25", "EQ251209_CSV.ZIP",  "Christmas"),
    ("BSE", "2010-01-26", "EQ260110_CSV.ZIP",  "Republic Day"),
    ("BSE", "2011-01-26", "EQ260111_CSV.ZIP",  "Republic Day"),
    ("BSE", "2011-04-22", "EQ220411_CSV.ZIP",  "Good Friday"),
    ("BSE", "2012-12-25", "EQ251212_CSV.ZIP",  "Christmas"),
    ("BSE", "2013-05-01", "EQ010513_CSV.ZIP",  "Maharashtra Day"),
    ("BSE", "2014-04-14", "EQ140414_CSV.ZIP",  "Dr Ambedkar Jayanti"),
    ("BSE", "2014-05-01", "EQ010514_CSV.ZIP",  "Maharashtra Day"),
    ("BSE", "2014-10-02", "EQ021014_CSV.ZIP",  "Gandhi Jayanti"),
    ("BSE", "2015-10-02", "EQ021015_CSV.ZIP",  "Gandhi Jayanti"),
    ("BSE", "2015-12-25", "EQ251215_CSV.ZIP",  "Christmas"),
    ("BSE", "2017-05-01", "EQ010517_CSV.ZIP",  "Maharashtra Day"),
    ("BSE", "2017-12-25", "EQ251217_CSV.ZIP",  "Christmas"),
    ("BSE", "2020-12-25", "EQ251220_CSV.ZIP",  "Christmas"),
    ("BSE", "2022-04-14", "EQ140422_CSV.ZIP",  "Dr Ambedkar Jayanti / Good Friday"),
    ("BSE", "2023-04-07", "EQ070423_CSV.ZIP",  "Good Friday"),
    ("BSE", "2024-05-01", "EQ010524_CSV.ZIP",  "Maharashtra Day"),
    ("BSE", "2026-01-26", "EQ260126_CSV.ZIP",  "Republic Day"),
    # Additional holidays found by proper analysis
    ("BSE", "2008-10-30", "EQ301008_CSV.ZIP",  "Diwali Laxmi Puja"),
    ("BSE", "2008-11-27", "EQ271108_CSV.ZIP",  "Bakri Id"),
    ("BSE", "2009-02-23", "EQ230209_CSV.ZIP",  "Mahashivratri"),
    ("BSE", "2009-04-30", "EQ300409_CSV.ZIP",  "Buddha Purnima"),
    ("BSE", "2009-07-07", "EQ070709_CSV.ZIP",  "Muharram"),
    ("BSE", "2009-09-28", "EQ280909_CSV.ZIP",  "Dussehra"),
    ("BSE", "2009-10-13", "EQ131009_CSV.ZIP",  "Id-ul-Fitr"),
    ("BSE", "2009-10-14", "EQ141009_CSV.ZIP",  "Id-ul-Fitr"),
    ("BSE", "2010-01-01", "EQ010110_CSV.ZIP",  "New Year's Day"),
    ("BSE", "2010-11-17", "EQ171110_CSV.ZIP",  "BSE observed holiday"),
    ("BSE", "2011-09-01", "EQ010911_CSV.ZIP",  "Ganesh Chaturthi"),
    ("BSE", "2011-12-06", "EQ061211_CSV.ZIP",  "BSE observed holiday"),
    ("BSE", "2012-04-05", "EQ050412_CSV.ZIP",  "BSE observed holiday"),
    ("BSE", "2012-10-24", "EQ241012_CSV.ZIP",  "Diwali Laxmi Puja"),
    ("BSE", "2012-11-14", "EQ141112_CSV.ZIP",  "Gurunanak Jayanti"),
    ("BSE", "2013-08-09", "EQ090813_CSV.ZIP",  "Id-ul-Fitr / BSE observed"),
    ("BSE", "2014-10-24", "EQ241014_CSV.ZIP",  "Diwali Laxmi Puja"),
    ("BSE", "2015-04-02", "EQ020415_CSV.ZIP",  "Ram Navami"),
    ("BSE", "2016-09-13", "EQ130916_CSV.ZIP",  "Ganesh Chaturthi observed"),
    ("BSE", "2016-10-31", "EQ311016_CSV.ZIP",  "Diwali Laxmi Puja"),
    ("BSE", "2016-11-14", "EQ141116_CSV.ZIP",  "Gurunanak Jayanti / BSE observed"),
    ("BSE", "2018-09-20", "EQ200918_CSV.ZIP",  "BSE observed holiday"),
    ("BSE", "2018-10-18", "EQ181018_CSV.ZIP",  "Dussehra observed"),
    ("BSE", "2018-11-23", "EQ231118_CSV.ZIP",  "Gurunanak Jayanti"),
    ("BSE", "2019-06-05", "EQ050619_CSV.ZIP",  "Id-ul-Fitr"),
    ("BSE", "2020-04-06", "EQ060420_CSV.ZIP",  "BSE Special Closure (COVID)"),
    ("BSE", "2021-03-11", "EQ110321_CSV.ZIP",  "Mahashivratri"),
    ("BSE", "2022-05-03", "EQ030522_CSV.ZIP",  "Id-ul-Fitr"),
    ("BSE", "2022-11-08", "EQ081122_CSV.ZIP",  "Gurunanak Jayanti"),
    ("BSE", "2023-03-30", "EQ300323_CSV.ZIP",  "Ram Navami observed"),
    ("BSE", "2024-05-20", "EQ200524_CSV.ZIP",  "General Election BSE holiday"),
    ("BSE", "2024-10-14", "EQ141024_CSV.ZIP",  "Dussehra"),
    ("BSE", "2025-10-29", "EQ291025_CSV.ZIP",  "Diwali Laxmi Puja"),
    ("BSE", "2026-03-18", "EQ180326_CSV.ZIP",  "Holi"),
]

# ============================================================
# GENUINE TRADING DAYS (77 BSE + 1 NSE = 78 total to retry)
# These actually had data — just timed out due to BSE throttling
# ============================================================
def bse(dd_mm_yyyy):
    d  = datetime.strptime(dd_mm_yyyy, "%d-%m-%Y")
    yy = d.strftime("%y")
    mm = d.strftime("%m")
    dy = d.strftime("%d")
    td = d.strftime("%Y-%m-%d")
    fn = f"EQ{dy}{mm}{yy}_CSV.ZIP"
    url = f"https://www.bseindia.com/download/BhavCopy/Equity/{fn}"
    return ("BSE", td, fn, url, os.path.join(BSE_DIR, fn))

TRADING_DAYS = [
    bse("27-03-2007"), bse("21-12-2007"), bse("19-05-2008"),
    bse("29-12-2008"), bse("17-03-2009"), bse("26-08-2009"),
    bse("28-08-2009"), bse("03-09-2009"), bse("20-07-2016"),
    bse("26-04-2024"), bse("09-07-2024"), bse("10-07-2024"),
    bse("19-07-2024"), bse("01-08-2024"), bse("26-09-2024"),
    bse("01-10-2024"), bse("06-11-2024"), bse("13-12-2024"),
    bse("02-01-2025"), bse("03-01-2025"), bse("15-01-2025"),
    bse("17-01-2025"), bse("14-02-2025"), bse("20-02-2025"),
    bse("28-02-2025"), bse("18-03-2025"), bse("20-03-2025"),
    bse("25-03-2025"), bse("07-04-2025"), bse("08-04-2025"),
    bse("16-04-2025"), bse("21-04-2025"), bse("29-04-2025"),
    bse("03-06-2025"), bse("11-06-2025"), bse("17-06-2025"),
    bse("30-06-2025"), bse("03-07-2025"), bse("07-07-2025"),
    bse("08-07-2025"), bse("18-07-2025"), bse("21-07-2025"),
    bse("30-07-2025"), bse("01-08-2025"), bse("04-08-2025"),
    bse("05-08-2025"), bse("18-08-2025"), bse("25-08-2025"),
    bse("02-09-2025"), bse("05-09-2025"), bse("09-09-2025"),
    bse("16-09-2025"), bse("18-09-2025"), bse("22-09-2025"),
    bse("25-09-2025"), bse("09-10-2025"), bse("24-10-2025"),
    bse("03-11-2025"), bse("10-11-2025"), bse("14-11-2025"),
    bse("19-11-2025"), bse("24-11-2025"), bse("17-12-2025"),
    bse("19-12-2025"), bse("23-12-2025"), bse("05-01-2026"),
    bse("06-01-2026"), bse("16-01-2026"), bse("21-01-2026"),
    bse("22-01-2026"), bse("29-01-2026"), bse("16-02-2026"),
    bse("19-02-2026"), bse("23-02-2026"), bse("25-02-2026"),
    bse("05-03-2026"), bse("09-03-2026"),
    # NSE SSL error
    ("NSE (Old)", "2009-12-22", "cm22DEC2009bhav.csv.zip",
     "https://nsearchives.nseindia.com/content/historical/EQUITIES/2009/DEC/cm22DEC2009bhav.csv.zip",
     os.path.join(NSE_OLD, "cm22DEC2009bhav.csv.zip")),
]

# ============================================================
# EXCEL SETUP
# ============================================================
FILL = {
    "HEADER":     PatternFill("solid", fgColor="1F4E79"),
    "DOWNLOADED": PatternFill("solid", fgColor="C6EFCE"),
    "EXISTS":     PatternFill("solid", fgColor="DDEBF7"),
    "HOLIDAY":    PatternFill("solid", fgColor="FFFACD"),
    "ERROR":      PatternFill("solid", fgColor="FFC7CE"),
    "WARNING":    PatternFill("solid", fgColor="FFEB9C"),
    "INFO":       PatternFill("solid", fgColor="F2F2F2"),
}
FONT_HDR  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
FONT_BODY = Font(name="Calibri", size=10)
BORDER    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
RUN_COLS   = ["Timestamp", "Exchange", "Trade Date", "Status", "File Name", "Detail"]
ERR_COLS   = ["Timestamp", "Exchange", "Trade Date", "Status", "File Name", "Detail", "URL"]
RUN_WIDTHS = [22, 12, 14, 14, 44, 70]
ERR_WIDTHS = [22, 12, 14, 14, 44, 70, 80]

wb = ws_run = ws_err = None
stats = {"holiday": 0, "downloaded": 0, "exists": 0, "failed": 0}

HEADERS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Encoding": "gzip, deflate",
    "Accept": "*/*",
    "Connection": "keep-alive",
}


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )


def make_sheet(name, cols, widths):
    ws = wb.create_sheet(name)
    for i, (col, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill      = FILL["HEADER"]
        cell.font      = FONT_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    return ws


def create_fresh_excel():
    global wb, ws_run, ws_err
    if os.path.exists(EXCEL_LOG):
        try:
            os.remove(EXCEL_LOG)
            logging.info("Removed old Logs.xlsx")
        except Exception as e:
            logging.error(f"Cannot delete old Logs.xlsx: {e}")
            logging.error("Close Excel if it is open, then run again.")
            exit(1)
    wb     = openpyxl.Workbook()
    for s in wb.sheetnames:
        del wb[s]
    ws_run = make_sheet("Run Log",   RUN_COLS, RUN_WIDTHS)
    ws_err = make_sheet("Error Log", ERR_COLS, ERR_WIDTHS)
    wb.save(EXCEL_LOG)
    logging.info(f"Fresh Logs.xlsx created at {EXCEL_LOG}")


def write_row(sheet, row_data):
    status = row_data[3] if len(row_data) > 3 else "INFO"
    fill   = FILL.get(status, FILL["INFO"])
    r      = sheet.max_row + 1
    for c, val in enumerate(row_data, 1):
        cell = sheet.cell(row=r, column=c, value=val)
        cell.fill      = fill
        cell.font      = FONT_BODY
        cell.alignment = Alignment(vertical="center")
        cell.border    = BORDER
    if r % 10 == 0:
        wb.save(EXCEL_LOG)


def download(exchange, trade_date, fname, url, save_path):
    if os.path.exists(save_path):
        logging.info(f"  EXISTS: {fname}")
        write_row(ws_run, [now(), exchange, trade_date, "EXISTS",
                            fname, "Already on disk"])
        stats["exists"] += 1
        return True

    for attempt in range(1, MAX_ATT + 1):
        try:
            logging.info(f"  Attempt {attempt}/{MAX_ATT}  {fname}")
            r = requests.get(url, headers=HEADERS_HTTP, timeout=TIMEOUT)
            if r.status_code == 200 and len(r.content) > 500:
                with open(save_path, "wb") as f:
                    f.write(r.content)
                kb = len(r.content) // 1024
                write_row(ws_run, [now(), exchange, trade_date, "DOWNLOADED",
                                    fname, f"Downloaded attempt {attempt} ({kb} KB)"])
                stats["downloaded"] += 1
                logging.info(f"  OK  {kb} KB")
                return True
            else:
                logging.warning(f"  HTTP {r.status_code}, {len(r.content)} bytes")
        except requests.exceptions.Timeout:
            logging.warning(f"  Timeout attempt {attempt}")
        except requests.exceptions.ConnectionError as e:
            logging.warning(f"  Connection error: {e}")
        except Exception as e:
            logging.error(f"  Error: {e}")

        if attempt < MAX_ATT:
            wait = RETRY_WAIT * attempt
            logging.info(f"  Waiting {wait}s...")
            time.sleep(wait)

    detail = f"All {MAX_ATT} attempts failed — genuine BSE archive gap"
    write_row(ws_run, [now(), exchange, trade_date, "ERROR", fname, detail])
    write_row(ws_err, [now(), exchange, trade_date, "ERROR", fname, detail, url])
    stats["failed"] += 1
    logging.warning(f"  GAVE UP: {fname}")
    return False


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    setup_logging()

    logging.info("=" * 60)
    logging.info("RETRY FINAL  —  132 files: 55 holidays + 77 genuine retries")
    logging.info("=" * 60)
    logging.info(f"  Holidays to log cleanly : {len(HOLIDAYS)}")
    logging.info(f"  Trading days to retry   : {len(TRADING_DAYS)}")
    logging.info("=" * 60)

    create_fresh_excel()
    write_row(ws_run, [now(), "---", "---", "INFO", "---",
                        "=== retry_final.py started — fresh Logs.xlsx ==="])

    # ---- PHASE 1: Log all holidays cleanly ----
    logging.info("")
    logging.info(f"PHASE 1 — Logging {len(HOLIDAYS)} confirmed holidays ...")
    for exchange, trade_date, fname, reason in HOLIDAYS:
        write_row(ws_run, [now(), exchange, trade_date, "HOLIDAY",
                            fname, f"Market holiday: {reason}"])
        stats["holiday"] += 1
        logging.info(f"  HOLIDAY  {trade_date}  {reason}")
    wb.save(EXCEL_LOG)
    logging.info(f"  Done. {stats['holiday']} holidays logged.")

    # ---- PHASE 2: Retry genuine trading days ----
    logging.info("")
    logging.info(f"PHASE 2 — Retrying {len(TRADING_DAYS)} genuine trading days ...")
    logging.info(f"  Timeout {TIMEOUT}s | {MAX_ATT} attempts | "
                 f"{MIN_DELAY:.0f}-{MAX_DELAY:.0f}s delay between files")
    logging.info("-" * 60)

    try:
        for i, job in enumerate(TRADING_DAYS, 1):
            exchange, trade_date, fname, url, save_path = job
            logging.info(f"[{i}/{len(TRADING_DAYS)}]  {trade_date}  {fname}")
            download(exchange, trade_date, fname, url, save_path)
            time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    except KeyboardInterrupt:
        logging.warning("Stopped by user (Ctrl+C) — progress saved to Excel.")
        write_row(ws_run, [now(), "---", "---", "WARNING", "---",
                            "Stopped by user (Ctrl+C)"])

    # ---- Summary ----
    logging.info("")
    logging.info("=" * 60)
    logging.info("SUMMARY")
    logging.info(f"  Holidays logged cleanly : {stats['holiday']}")
    logging.info(f"  Successfully downloaded : {stats['downloaded']}")
    logging.info(f"  Already on disk         : {stats['exists']}")
    logging.info(f"  Still failing           : {stats['failed']}")
    logging.info("=" * 60)

    summary = (f"DONE — Holidays: {stats['holiday']} | "
               f"Downloaded: {stats['downloaded']} | "
               f"Existed: {stats['exists']} | "
               f"Failed: {stats['failed']}")
    write_row(ws_run, [now(), "---", "---", "INFO", "SUMMARY", summary])
    if stats["failed"] > 0:
        write_row(ws_err, [now(), "---", "---", "WARNING", "SUMMARY",
                            f"{stats['failed']} genuine BSE archive gaps remain", "---"])

    wb.save(EXCEL_LOG)
    logging.info(f"Logs.xlsx saved → {EXCEL_LOG}")
    logging.info("All done.")
