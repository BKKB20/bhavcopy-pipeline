import requests
import os
import time
import random
import logging
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# FOLDER PATHS
# ============================================================
NSE_OLD_DIR  = r"D:\Bhavya Khaitan\Stocks\StockData\NSE\Old Format"
NSE_NEW_DIR  = r"D:\Bhavya Khaitan\Stocks\StockData\NSE\New Format"
BSE_DIR      = r"D:\Bhavya Khaitan\Stocks\StockData\BSE"
EXCEL_LOG    = r"D:\Bhavya Khaitan\Stocks\StockData\Logs.xlsx"

# ============================================================
# DATE RANGES
# ============================================================
NSE_OLD_START = date(2000, 1,  1)
NSE_OLD_END   = date(2024, 7,  7)
NSE_NEW_START = date(2024, 7,  8)
NSE_NEW_END   = date.today()
BSE_START     = date(2007, 1,  1)
BSE_END       = date.today()

# ============================================================
# DELAY SETTINGS
# ============================================================
MIN_DELAY = 1.5
MAX_DELAY = 3.0

# ============================================================
# COUNTERS
# ============================================================
stats = {
    "downloaded":      0,
    "skipped_exists":  0,
    "skipped_holiday": 0,
    "errors":          0,
}

# ============================================================
# EXCEL LOG SETUP
# Two sheets inside Logs.xlsx:
#   "Run Log"   — every event with timestamp, exchange, date, status, detail
#   "Error Log" — only warnings and errors
# If the file already exists, new rows are APPENDED (not overwritten)
# ============================================================

# Colour fills
FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")
FILL_OK       = PatternFill("solid", fgColor="C6EFCE")
FILL_EXISTS   = PatternFill("solid", fgColor="DDEBF7")
FILL_HOLIDAY  = PatternFill("solid", fgColor="FFFACD")
FILL_ERROR    = PatternFill("solid", fgColor="FFC7CE")
FILL_WARNING  = PatternFill("solid", fgColor="FFEB9C")
FILL_SUMMARY  = PatternFill("solid", fgColor="F2F2F2")

FONT_HEADER   = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
FONT_NORMAL   = Font(name="Calibri", size=10)

THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

RUN_HEADERS  = ["Timestamp", "Exchange", "Trade Date", "Status", "File Name", "Detail"]
ERR_HEADERS  = ["Timestamp", "Exchange", "Trade Date", "Status", "File Name", "Detail", "URL"]

RUN_COL_WIDTHS = [22, 12, 14, 14, 44, 60]
ERR_COL_WIDTHS = [22, 12, 14, 14, 44, 60, 80]

wb      = None
ws_run  = None
ws_err  = None


def _style_header_row(ws, headers, col_widths):
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill      = FILL_HEADER
        cell.font      = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def setup_excel_log():
    global wb, ws_run, ws_err

    if os.path.exists(EXCEL_LOG):
        wb = openpyxl.load_workbook(EXCEL_LOG)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Run Log" not in wb.sheetnames:
        ws_run = wb.create_sheet("Run Log")
        _style_header_row(ws_run, RUN_HEADERS, RUN_COL_WIDTHS)
    else:
        ws_run = wb["Run Log"]

    if "Error Log" not in wb.sheetnames:
        ws_err = wb.create_sheet("Error Log")
        _style_header_row(ws_err, ERR_HEADERS, ERR_COL_WIDTHS)
    else:
        ws_err = wb["Error Log"]

    wb.save(EXCEL_LOG)
    logging.info(f"Excel log ready -> {EXCEL_LOG}")


def _row_fill(status):
    return {
        "DOWNLOADED": FILL_OK,
        "EXISTS":     FILL_EXISTS,
        "HOLIDAY":    FILL_HOLIDAY,
        "ERROR":      FILL_ERROR,
        "WARNING":    FILL_WARNING,
        "INFO":       FILL_SUMMARY,
    }.get(status, FILL_SUMMARY)


def log_to_excel(sheet, row_data):
    status  = row_data[3] if len(row_data) > 3 else ""
    fill    = _row_fill(status)
    row_num = sheet.max_row + 1
    for col_idx, value in enumerate(row_data, start=1):
        cell = sheet.cell(row=row_num, column=col_idx, value=value)
        cell.fill      = fill
        cell.font      = FONT_NORMAL
        cell.alignment = Alignment(vertical="center")
        cell.border    = THIN_BORDER
    if row_num % 50 == 0:
        wb.save(EXCEL_LOG)


# ============================================================
# CONSOLE LOGGING
# ============================================================
def setup_console_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )


# ============================================================
# HELPERS
# ============================================================
MONTHS = ["JAN","FEB","MAR","APR","MAY","JUN",
          "JUL","AUG","SEP","OCT","NOV","DEC"]

HEADERS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Encoding": "gzip, deflate",
    "Accept": "*/*",
    "Connection": "keep-alive",
}


def make_dirs():
    for d in [NSE_OLD_DIR, NSE_NEW_DIR, BSE_DIR,
              os.path.dirname(EXCEL_LOG)]:
        os.makedirs(d, exist_ok=True)


def all_weekdays(start, end):
    current = start
    while current <= end:
        if current.weekday() < 5:
            yield current
        current += timedelta(days=1)


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def download_file(url, save_path, label, exchange, trade_date):
    fname = os.path.basename(save_path)

    if os.path.exists(save_path):
        stats["skipped_exists"] += 1
        log_to_excel(ws_run, [now_str(), exchange, trade_date, "EXISTS", fname,
                               "Already downloaded, skipped"])
        return "exists"

    try:
        resp = requests.get(url, headers=HEADERS_HTTP, timeout=30)

        if resp.status_code == 200 and len(resp.content) > 500:
            with open(save_path, "wb") as f:
                f.write(resp.content)
            size_kb = len(resp.content) // 1024
            detail  = f"Downloaded successfully ({size_kb} KB)"
            logging.info(f"OK  {label}  {size_kb} KB")
            stats["downloaded"] += 1
            log_to_excel(ws_run, [now_str(), exchange, trade_date, "DOWNLOADED", fname, detail])
            return "ok"

        else:
            detail = f"HTTP {resp.status_code} — market holiday or non-trading day"
            stats["skipped_holiday"] += 1
            log_to_excel(ws_run, [now_str(), exchange, trade_date, "HOLIDAY", fname, detail])
            return "holiday"

    except requests.exceptions.Timeout:
        detail = "Request timed out after 30 seconds"
        logging.warning(f"TIMEOUT  {label}")
        stats["errors"] += 1
        log_to_excel(ws_run, [now_str(), exchange, trade_date, "ERROR", fname, detail])
        log_to_excel(ws_err, [now_str(), exchange, trade_date, "ERROR", fname, detail, url])
        return "error"

    except requests.exceptions.ConnectionError as e:
        detail = f"Connection error: {e}"
        logging.warning(f"CONN_ERR  {label}")
        stats["errors"] += 1
        log_to_excel(ws_run, [now_str(), exchange, trade_date, "ERROR", fname, detail])
        log_to_excel(ws_err, [now_str(), exchange, trade_date, "ERROR", fname, detail, url])
        return "error"

    except Exception as e:
        detail = f"Unexpected error: {e}"
        logging.error(f"ERROR  {label}  {e}")
        stats["errors"] += 1
        log_to_excel(ws_run, [now_str(), exchange, trade_date, "ERROR", fname, detail])
        log_to_excel(ws_err, [now_str(), exchange, trade_date, "ERROR", fname, detail, url])
        return "error"


def polite_sleep():
    time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))


# ============================================================
# NSE OLD FORMAT  (Jan 2000 - Jul 7 2024)
# ============================================================
def download_nse_old():
    logging.info("--- NSE OLD FORMAT (Jan 2000 - Jul 7, 2024) ---")
    log_to_excel(ws_run, [now_str(), "---", "---", "INFO", "---",
                           "=== NSE OLD FORMAT started (Jan 2000 - Jul 7 2024) ==="])

    session = requests.Session()
    try:
        session.get("https://www.nseindia.com", headers=HEADERS_HTTP, timeout=15)
        logging.info("NSE session cookie obtained.")
    except Exception as e:
        logging.warning(f"Could not get NSE cookie: {e}")
        log_to_excel(ws_err, [now_str(), "NSE", "---", "WARNING", "---",
                               f"Could not pre-fetch NSE session cookie: {e}",
                               "https://www.nseindia.com"])

    for d in all_weekdays(NSE_OLD_START, NSE_OLD_END):
        dd   = d.strftime("%d")
        mon  = MONTHS[d.month - 1]
        yyyy = d.strftime("%Y")
        fname     = f"cm{dd}{mon}{yyyy}bhav.csv.zip"
        url       = (
            f"https://nsearchives.nseindia.com/content/historical/"
            f"EQUITIES/{yyyy}/{mon}/{fname}"
        )
        save_path = os.path.join(NSE_OLD_DIR, fname)
        download_file(url, save_path, f"NSE-OLD {d}", "NSE (Old)", str(d))
        polite_sleep()


# ============================================================
# NSE NEW UDiFF FORMAT  (Jul 8 2024 - today)
# ============================================================
def download_nse_new():
    logging.info("--- NSE NEW UDiFF FORMAT (Jul 8, 2024 - today) ---")
    log_to_excel(ws_run, [now_str(), "---", "---", "INFO", "---",
                           "=== NSE NEW UDiFF FORMAT started (Jul 8 2024 - today) ==="])

    for d in all_weekdays(NSE_NEW_START, NSE_NEW_END):
        date_str  = d.strftime("%Y%m%d")
        fname     = f"BhavCopy_NSE_CM_0_0_0_{date_str}_F_0000.csv.zip"
        url       = f"https://nsearchives.nseindia.com/content/cm/{fname}"
        save_path = os.path.join(NSE_NEW_DIR, fname)
        download_file(url, save_path, f"NSE-NEW {d}", "NSE (New)", str(d))
        polite_sleep()


# ============================================================
# BSE  (Jan 2007 - today)
# ============================================================
def download_bse():
    logging.info("--- BSE (Jan 2007 - today) ---")
    log_to_excel(ws_run, [now_str(), "---", "---", "INFO", "---",
                           "=== BSE started (Jan 2007 - today) ==="])

    for d in all_weekdays(BSE_START, BSE_END):
        dd        = d.strftime("%d")
        mm        = d.strftime("%m")
        yy        = d.strftime("%y")
        fname     = f"EQ{dd}{mm}{yy}_CSV.ZIP"
        url       = (
            f"https://www.bseindia.com/download/BhavCopy/Equity/{fname}"
        )
        save_path = os.path.join(BSE_DIR, fname)
        download_file(url, save_path, f"BSE {d}", "BSE", str(d))
        polite_sleep()


# ============================================================
# SUMMARY
# ============================================================
def write_summary():
    lines = [
        f"Freshly downloaded : {stats['downloaded']}",
        f"Already existed    : {stats['skipped_exists']}",
        f"Holidays / no data : {stats['skipped_holiday']}",
        f"Errors             : {stats['errors']}",
    ]
    for line in lines:
        logging.info(line)
        log_to_excel(ws_run, [now_str(), "---", "---", "INFO", "SUMMARY", line])

    if stats["errors"] > 0:
        note = f"{stats['errors']} failures — check Error Log sheet in Logs.xlsx"
        logging.warning(note)
        log_to_excel(ws_err, [now_str(), "---", "---", "WARNING", "SUMMARY", note, "---"])


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    setup_console_logging()
    make_dirs()
    setup_excel_log()

    logging.info("BHAVCOPY BULK DOWNLOADER started")
    logging.info("Safe to stop with Ctrl+C and restart anytime.")
    log_to_excel(ws_run, [now_str(), "---", "---", "INFO", "---",
                           "BHAVCOPY BULK DOWNLOADER started"])

    try:
        download_nse_old()
        download_nse_new()
        download_bse()

    except KeyboardInterrupt:
        logging.warning("Stopped by user (Ctrl+C).")
        log_to_excel(ws_run, [now_str(), "---", "---", "WARNING", "---",
                               "Stopped by user (Ctrl+C)"])
        log_to_excel(ws_err, [now_str(), "---", "---", "WARNING", "---",
                               "Stopped by user (Ctrl+C)", "---"])

    finally:
        write_summary()
        wb.save(EXCEL_LOG)
        logging.info(f"Excel log saved -> {EXCEL_LOG}")
        logging.info("Done.")
