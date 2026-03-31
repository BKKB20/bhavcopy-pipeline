"""
NSE Bhav Copy ZIP Extractor
----------------------------
Extracts the single CSV inside each zip file in the source folder.
Logs execution summary and any errors to an Excel file.

Save this file at:
    C:\\Users\\Bhavya Khaitan\\BK\\Stocks\\StockData\\NSE\\extract_nse_zips.py

Run from Command Prompt:
    cd C:\\Users\\Bhavya Khaitan\\BK\\Stocks\\StockData\\NSE
    python extract_nse_zips.py

Requires openpyxl:
    pip install openpyxl
"""

import zipfile
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── CONFIGURATION ─────────────────────────────────────────────────────────────

SOURCE_DIR = r"C:\Users\Bhavya Khaitan\BK\Stocks\StockData\NSE\Old Format"
OUTPUT_DIR = r"C:\Users\Bhavya Khaitan\BK\Stocks\StockData\NSE\Old Format\Extracted"
LOG_FILE   = r"C:\Users\Bhavya Khaitan\BK\Stocks\StockData\NSE\log sheet old format.xlsx"

# ──────────────────────────────────────────────────────────────────────────────


# ── COLOUR PALETTE ────────────────────────────────────────────────────────────
CLR_HEADER_EXEC  = "1F4E79"   # dark blue   – execution log header
CLR_HEADER_ERR   = "7B0000"   # dark red    – error log header
CLR_ROW_SUCCESS  = "E2EFDA"   # light green – extracted
CLR_ROW_SKIP     = "FFF2CC"   # yellow      – skipped
CLR_ROW_ERROR    = "FCE4D6"   # orange/red  – error
CLR_SUMMARY_HDR  = "2E75B6"   # medium blue – summary title
CLR_WHITE        = "FFFFFF"
CLR_LIGHT_BLUE   = "DDEBF7"
# ──────────────────────────────────────────────────────────────────────────────


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def write_excel_log(log_path, run_start, run_end,
                    total, success, skipped, error_rows, execution_rows):
    """Create the Excel log workbook with three sheets."""

    wb = openpyxl.Workbook()

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 1 – Execution Summary
    # ──────────────────────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Execution Summary"
    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 62

    # Title
    ws_sum.merge_cells("A1:B1")
    t = ws_sum["A1"]
    t.value     = "NSE Bhav Copy Extraction – Run Summary"
    t.font      = Font(bold=True, color=CLR_WHITE, size=13)
    t.fill      = PatternFill("solid", fgColor=CLR_SUMMARY_HDR)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    rows = [
        ("Run Start",                         run_start.strftime("%Y-%m-%d %H:%M:%S")),
        ("Run End",                            run_end.strftime("%Y-%m-%d %H:%M:%S")),
        ("Duration (seconds)",                 round((run_end - run_start).total_seconds(), 1)),
        ("Source Folder",                      SOURCE_DIR),
        ("Output Folder",                      OUTPUT_DIR),
        ("Total ZIP Files Found",              total),
        ("Successfully Extracted",             success),
        ("Skipped (already existed / empty)",  skipped),
        ("Errors",                             len(error_rows)),
    ]
    for r, (label, value) in enumerate(rows, 2):
        lc = ws_sum.cell(row=r, column=1, value=label)
        vc = ws_sum.cell(row=r, column=2, value=value)
        bg = CLR_LIGHT_BLUE if r % 2 == 0 else CLR_WHITE
        for c in (lc, vc):
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = thin_border()
            c.alignment = Alignment(vertical="center")
        lc.font = Font(bold=True)

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 2 – Execution Log (every file)
    # ──────────────────────────────────────────────────────────────────────────
    ws_exec = wb.create_sheet("Execution Log")

    exec_cols = [
        ("#",              6),
        ("ZIP File Name", 46),
        ("CSV Extracted", 46),
        ("Status",        14),
        ("Timestamp",     22),
    ]
    for ci, (hdr, w) in enumerate(exec_cols, 1):
        c = ws_exec.cell(row=1, column=ci, value=hdr)
        c.font      = Font(bold=True, color=CLR_WHITE, size=11)
        c.fill      = PatternFill("solid", fgColor=CLR_HEADER_EXEC)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
        ws_exec.column_dimensions[get_column_letter(ci)].width = w
    ws_exec.row_dimensions[1].height = 20

    status_clr = {"Extracted": CLR_ROW_SUCCESS,
                  "Skipped":   CLR_ROW_SKIP,
                  "Error":     CLR_ROW_ERROR}

    for ri, row in enumerate(execution_rows, 2):
        vals = [row["num"], row["zip_name"],
                row.get("csv_name", ""), row["status"], row["timestamp"]]
        bg = status_clr.get(row["status"], CLR_WHITE)
        for ci, val in enumerate(vals, 1):
            c = ws_exec.cell(row=ri, column=ci, value=val)
            c.fill   = PatternFill("solid", fgColor=bg)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center")

    ws_exec.freeze_panes = "A2"
    ws_exec.auto_filter.ref = f"A1:E{len(execution_rows) + 1}"

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 3 – Error Log
    # ──────────────────────────────────────────────────────────────────────────
    ws_err = wb.create_sheet("Error Log")

    err_cols = [
        ("#",              6),
        ("ZIP File Name", 46),
        ("Error Type",    22),
        ("Error Detail",  62),
        ("Timestamp",     22),
    ]
    for ci, (hdr, w) in enumerate(err_cols, 1):
        c = ws_err.cell(row=1, column=ci, value=hdr)
        c.font      = Font(bold=True, color=CLR_WHITE, size=11)
        c.fill      = PatternFill("solid", fgColor=CLR_HEADER_ERR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
        ws_err.column_dimensions[get_column_letter(ci)].width = w
    ws_err.row_dimensions[1].height = 20

    if error_rows:
        for ri, row in enumerate(error_rows, 2):
            vals = [row["num"], row["zip_name"],
                    row["error_type"], row["error_detail"], row["timestamp"]]
            for ci, val in enumerate(vals, 1):
                c = ws_err.cell(row=ri, column=ci, value=val)
                c.fill      = PatternFill("solid", fgColor=CLR_ROW_ERROR)
                c.border    = thin_border()
                c.alignment = Alignment(vertical="center", wrap_text=True)
        ws_err.freeze_panes = "A2"
        ws_err.auto_filter.ref = f"A1:E{len(error_rows) + 1}"
    else:
        ws_err.merge_cells("A2:E2")
        nc = ws_err["A2"]
        nc.value     = "No errors encountered during this run."
        nc.font      = Font(bold=True, color="375623")
        nc.fill      = PatternFill("solid", fgColor=CLR_ROW_SUCCESS)
        nc.alignment = Alignment(horizontal="center", vertical="center")
        ws_err.row_dimensions[2].height = 22

    wb.save(log_path)
    print(f"  Log saved → {log_path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def extract_all():
    source = Path(SOURCE_DIR)
    output = Path(OUTPUT_DIR)
    output.mkdir(parents=True, exist_ok=True)

    zip_files = sorted(source.glob("*.zip"))
    total     = len(zip_files)
    print(f"Found {total} zip files in:\n  {source}\n")

    run_start      = datetime.now()
    success        = 0
    skipped        = 0
    error_rows     = []
    execution_rows = []

    for i, zip_path in enumerate(zip_files, 1):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                names = zf.namelist()

                if not names:           # empty zip
                    skipped += 1
                    execution_rows.append({"num": i, "zip_name": zip_path.name,
                                           "csv_name": "", "status": "Skipped",
                                           "timestamp": ts})
                    continue

                csv_name  = names[0]
                dest_file = output / csv_name

                if dest_file.exists():  # already extracted
                    skipped += 1
                    execution_rows.append({"num": i, "zip_name": zip_path.name,
                                           "csv_name": csv_name, "status": "Skipped",
                                           "timestamp": ts})
                else:
                    zf.extract(csv_name, output)
                    success += 1
                    execution_rows.append({"num": i, "zip_name": zip_path.name,
                                           "csv_name": csv_name, "status": "Extracted",
                                           "timestamp": ts})

        except zipfile.BadZipFile:
            print(f"  [ERROR] Bad zip: {zip_path.name}")
            error_rows.append({"num": i, "zip_name": zip_path.name,
                                "error_type": "BadZipFile",
                                "error_detail": "Corrupt or not a valid ZIP archive.",
                                "timestamp": ts})
            execution_rows.append({"num": i, "zip_name": zip_path.name,
                                    "csv_name": "", "status": "Error", "timestamp": ts})

        except Exception as e:
            print(f"  [ERROR] {zip_path.name}: {e}")
            error_rows.append({"num": i, "zip_name": zip_path.name,
                                "error_type": type(e).__name__,
                                "error_detail": str(e),
                                "timestamp": ts})
            execution_rows.append({"num": i, "zip_name": zip_path.name,
                                    "csv_name": "", "status": "Error", "timestamp": ts})

        if i % 500 == 0:
            print(f"  [{i}/{total}] processed ...")

    run_end = datetime.now()

    print(f"\n── Results ──────────────────────────────────")
    print(f"  Extracted : {success}")
    print(f"  Skipped   : {skipped}  (already existed or empty)")
    print(f"  Errors    : {len(error_rows)}")
    print(f"  Duration  : {round((run_end - run_start).total_seconds(), 1)} seconds")
    print(f"─────────────────────────────────────────────")

    print("\nWriting Excel log ...")
    write_excel_log(LOG_FILE, run_start, run_end,
                    total, success, skipped, error_rows, execution_rows)

    print(f"\nCSV files are in:\n  {output}")
    print("\nAll done!")


if __name__ == "__main__":
    extract_all()
