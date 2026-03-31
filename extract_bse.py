"""
BSE Bhav Copy ZIP Extractor
-----------------------------
Extracts CSVs from the BSE folder.
Appends new sheets to the existing Excel log workbook
(renamed to "log sheet.xlsx").

Save this file at:
    C:\\Users\\Bhavya Khaitan\\BK\\Stocks\\StockData\\BSE\\extract_bse.py

Run from Command Prompt:
    cd C:\\Users\\Bhavya Khaitan\\BK\\Stocks\\StockData\\BSE
    python extract_bse.py

Requires openpyxl:
    pip install openpyxl

IMPORTANT:
    1. Rename your existing log file first:
       "log sheet nse.xlsx"  →  "log sheet.xlsx"
       (Move it to: C:\\Users\\Bhavya Khaitan\\BK\\Stocks\\StockData\\)
       OR keep it in NSE folder and update LOG_FILE path below.
    2. This script will ADD three new sheets to that workbook:
       - "BSE Execution Summary"
       - "BSE Execution Log"
       - "BSE Error Log"
       It will NOT touch any existing sheets.
"""

import zipfile
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── CONFIGURATION ─────────────────────────────────────────────────────────────

SOURCE_DIR = r"C:\Users\Bhavya Khaitan\BK\Stocks\StockData\BSE"
OUTPUT_DIR = r"C:\Users\Bhavya Khaitan\BK\Stocks\StockData\BSE\Extracted"

# Update this path to wherever you placed/renamed the log file
LOG_FILE   = r"C:\Users\Bhavya Khaitan\BK\Stocks\StockData\log sheet.xlsx"

# ──────────────────────────────────────────────────────────────────────────────


# ── COLOUR PALETTE ────────────────────────────────────────────────────────────
CLR_HEADER_EXEC  = "833C00"   # dark orange  – BSE execution log header
CLR_HEADER_ERR   = "7B0000"   # dark red     – error log header
CLR_ROW_SUCCESS  = "E2EFDA"   # light green  – extracted
CLR_ROW_SKIP     = "FFF2CC"   # yellow       – skipped
CLR_ROW_ERROR    = "FCE4D6"   # orange/red   – error
CLR_SUMMARY_HDR  = "C55A11"   # orange       – BSE summary header
CLR_WHITE        = "FFFFFF"
CLR_LIGHT_ORANGE = "FCE4D6"
CLR_LIGHT_BLUE   = "DDEBF7"
# ──────────────────────────────────────────────────────────────────────────────


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def add_bse_sheets(log_path, run_start, run_end,
                   total, success, skipped, error_rows, execution_rows):
    """Open the existing workbook (or create new) and append BSE sheets."""

    if Path(log_path).exists():
        wb = openpyxl.load_workbook(log_path)
        print(f"  Opened existing log: {log_path}")
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        print(f"  Creating new log: {log_path}")

    # Remove BSE sheets if they already exist (re-run protection)
    for sheet_name in ["BSE Execution Summary", "BSE Execution Log", "BSE Error Log"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    # ── Sheet : BSE Execution Summary ────────────────────────────────────────
    ws_sum = wb.create_sheet("BSE Execution Summary")
    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 62

    ws_sum.merge_cells("A1:B1")
    t = ws_sum["A1"]
    t.value     = "BSE Bhav Copy Extraction – Run Summary"
    t.font      = Font(bold=True, color=CLR_WHITE, size=13)
    t.fill      = PatternFill("solid", fgColor=CLR_SUMMARY_HDR)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    summary_rows = [
        ("Run Start",                         run_start.strftime("%Y-%m-%d %H:%M:%S")),
        ("Run End",                            run_end.strftime("%Y-%m-%d %H:%M:%S")),
        ("Duration (seconds)",                 round((run_end - run_start).total_seconds(), 1)),
        ("Source Folder",                      SOURCE_DIR),
        ("Output Folder",                      OUTPUT_DIR),
        ("Total ZIP Files Found",              total),
        ("Successfully Extracted",             success),
        ("Skipped (already existed / empty)",  skipped),
        ("Errors",                             len(error_rows)),
    ]
    for r, (label, value) in enumerate(summary_rows, 2):
        lc = ws_sum.cell(row=r, column=1, value=label)
        vc = ws_sum.cell(row=r, column=2, value=value)
        bg = CLR_LIGHT_ORANGE if r % 2 == 0 else CLR_WHITE
        for c in (lc, vc):
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = thin_border()
            c.alignment = Alignment(vertical="center")
        lc.font = Font(bold=True)

    # ── Sheet : BSE Execution Log ─────────────────────────────────────────────
    ws_exec = wb.create_sheet("BSE Execution Log")

    exec_cols = [
        ("#",              6),
        ("ZIP File Name", 30),
        ("CSV Extracted", 30),
        ("Status",        14),
        ("Timestamp",     22),
    ]
    for ci, (hdr, w) in enumerate(exec_cols, 1):
        c = ws_exec.cell(row=1, column=ci, value=hdr)
        c.font      = Font(bold=True, color=CLR_WHITE, size=11)
        c.fill      = PatternFill("solid", fgColor=CLR_HEADER_EXEC)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
        ws_exec.column_dimensions[get_column_letter(ci)].width = w
    ws_exec.row_dimensions[1].height = 20

    status_clr = {
        "Extracted": CLR_ROW_SUCCESS,
        "Skipped":   CLR_ROW_SKIP,
        "Error":     CLR_ROW_ERROR,
    }

    for ri, row in enumerate(execution_rows, 2):
        vals = [row["num"], row["zip_name"],
                row.get("csv_name", ""), row["status"], row["timestamp"]]
        bg = status_clr.get(row["status"], CLR_WHITE)
        for ci, val in enumerate(vals, 1):
            c = ws_exec.cell(row=ri, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = thin_border()
            c.alignment = Alignment(vertical="center")

    ws_exec.freeze_panes = "A2"
    ws_exec.auto_filter.ref = f"A1:E{len(execution_rows) + 1}"

    # ── Sheet : BSE Error Log ─────────────────────────────────────────────────
    ws_err = wb.create_sheet("BSE Error Log")

    err_cols = [
        ("#",              6),
        ("ZIP File Name", 30),
        ("Error Type",    22),
        ("Error Detail",  62),
        ("Timestamp",     22),
    ]
    for ci, (hdr, w) in enumerate(err_cols, 1):
        c = ws_err.cell(row=1, column=ci, value=hdr)
        c.font      = Font(bold=True, color=CLR_WHITE, size=11)
        c.fill      = PatternFill("solid", fgColor=CLR_HEADER_ERR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
        ws_err.column_dimensions[get_column_letter(ci)].width = w
    ws_err.row_dimensions[1].height = 20

    if error_rows:
        for ri, row in enumerate(error_rows, 2):
            vals = [row["num"], row["zip_name"],
                    row["error_type"], row["error_detail"], row["timestamp"]]
            for ci, val in enumerate(vals, 1):
                c = ws_err.cell(row=ri, column=ci, value=val)
                c.fill      = PatternFill("solid", fgColor=CLR_ROW_ERROR)
                c.border    = thin_border()
                c.alignment = Alignment(vertical="center", wrap_text=True)
        ws_err.freeze_panes = "A2"
        ws_err.auto_filter.ref = f"A1:E{len(error_rows) + 1}"
    else:
        ws_err.merge_cells("A2:E2")
        nc = ws_err["A2"]
        nc.value     = "No errors encountered during this run."
        nc.font      = Font(bold=True, color="375623")
        nc.fill      = PatternFill("solid", fgColor=CLR_ROW_SUCCESS)
        nc.alignment = Alignment(horizontal="center", vertical="center")
        ws_err.row_dimensions[2].height = 22

    wb.save(log_path)
    print(f"  Log saved → {log_path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def extract_all():
    source = Path(SOURCE_DIR)
    output = Path(OUTPUT_DIR)
    output.mkdir(parents=True, exist_ok=True)

    zip_files = sorted(source.glob("*.ZIP")) + sorted(source.glob("*.zip"))
    # Deduplicate in case of mixed case on some systems
    seen = set()
    unique_zips = []
    for z in zip_files:
        if z.name.upper() not in seen:
            seen.add(z.name.upper())
            unique_zips.append(z)
    zip_files = sorted(unique_zips, key=lambda x: x.name.upper())

    total     = len(zip_files)
    print(f"Found {total} ZIP files in:\n  {source}\n")

    run_start      = datetime.now()
    success        = 0
    skipped        = 0
    error_rows     = []
    execution_rows = []

    for i, zip_path in enumerate(zip_files, 1):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                names = zf.namelist()

                if not names:
                    skipped += 1
                    execution_rows.append({
                        "num": i, "zip_name": zip_path.name,
                        "csv_name": "", "status": "Skipped", "timestamp": ts
                    })
                    continue

                csv_name  = names[0]
                dest_file = output / csv_name

                if dest_file.exists():
                    skipped += 1
                    execution_rows.append({
                        "num": i, "zip_name": zip_path.name,
                        "csv_name": csv_name, "status": "Skipped", "timestamp": ts
                    })
                else:
                    zf.extract(csv_name, output)
                    success += 1
                    execution_rows.append({
                        "num": i, "zip_name": zip_path.name,
                        "csv_name": csv_name, "status": "Extracted", "timestamp": ts
                    })

        except zipfile.BadZipFile:
            print(f"  [ERROR] Bad zip: {zip_path.name}")
            error_rows.append({
                "num": i, "zip_name": zip_path.name,
                "error_type": "BadZipFile",
                "error_detail": "Corrupt or not a valid ZIP archive.",
                "timestamp": ts
            })
            execution_rows.append({
                "num": i, "zip_name": zip_path.name,
                "csv_name": "", "status": "Error", "timestamp": ts
            })

        except Exception as e:
            print(f"  [ERROR] {zip_path.name}: {e}")
            error_rows.append({
                "num": i, "zip_name": zip_path.name,
                "error_type": type(e).__name__,
                "error_detail": str(e),
                "timestamp": ts
            })
            execution_rows.append({
                "num": i, "zip_name": zip_path.name,
                "csv_name": "", "status": "Error", "timestamp": ts
            })

        if i % 500 == 0:
            print(f"  [{i}/{total}] processed ...")

    run_end = datetime.now()

    print(f"\n── Results ──────────────────────────────────")
    print(f"  Extracted : {success}")
    print(f"  Skipped   : {skipped}  (already existed or empty)")
    print(f"  Errors    : {len(error_rows)}")
    print(f"  Duration  : {round((run_end - run_start).total_seconds(), 1)} seconds")
    print(f"─────────────────────────────────────────────")

    print("\nWriting to Excel log ...")
    add_bse_sheets(
        LOG_FILE, run_start, run_end,
        total, success, skipped, error_rows, execution_rows
    )

    print(f"\nCSV files are in:\n  {output}")
    print("\nAll done!")


if __name__ == "__main__":
    extract_all()
